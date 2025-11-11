import gradio as gr
import pandas as pd
import numpy as np
import openpyxl
import io
import random
import logging
from scipy.signal import medfilt
from scipy.ndimage import gaussian_filter1d
import matplotlib.pyplot as plt
import tempfile
import os

# --- Configuración de Logging ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- Constantes Globales ---
ARCHIVOS_CON_LIMITE = ["1. OQ_MAPEO.xlsm", "4. PQ_RUTA_20.xlsm", "5. PQ_RUTA_80.xlsm"]
HOJAS_A_IGNORAR = ["CONSOLIDADO", "GRAFICOS", "RESUMEN", "TABLA", "RESULTADOS", "SUMMARY", "GRAFICO"] 

# --- CONFIGURACIÓN BASE (Presets) ---
CONFIGURACION_BASE = {
    "OQ Mapeo": {
        "archivo": "1. OQ MAPEO 72 INV.xlsm",        
        "variacion_min": 0.01, "variacion_max": 0.02,
        "amplitud": 0.30, "sigma": 12, "punto_pico": 0.5,
        "offset_min": -0.5, "offset_max": 0.0,
        "prob_limpieza_picos": 0.9
    },
    "OQ Apertura": {
        "archivo": "2. OQ APERTURA 72 INV.xlsm",     
        "variacion_min": 0.03, "variacion_max": 0.05,
        "amplitud": 0.40, "sigma": 8, "punto_pico": 0.6,
        "offset_min": -1.0, "offset_max": -0.2,
        "prob_limpieza_picos": 0.5
    },
    "OQ Apagado": {
        "archivo": "3. OQ APAGADO 72 INV.xlsm",      
        "variacion_min": 0.01, "variacion_max": 0.02,
        "amplitud": 0.50, "sigma": 20, "punto_pico": 0.4,
        "offset_min": -1.2, "offset_max": -0.3,
        "prob_limpieza_picos": 0.9
    },
    "PQ Ruta 20%": {
        "archivo": "4. PQ RUTA 20 72 INV.xlsm",      
        "variacion_min": 0.02, "variacion_max": 0.04,
        "amplitud": 0.35, "sigma": 12, "punto_pico": 0.5,
        "offset_min": -0.9, "offset_max": -0.2,
        "prob_limpieza_picos": 0.7
    },
    "PQ Ruta 80%": {
        "archivo": "5. PQ RUTA 80 72 INV.xlsm",      
        "variacion_min": 0.02, "variacion_max": 0.04,
        "amplitud": 0.35, "sigma": 12, "punto_pico": 0.5,
        "offset_min": -0.9, "offset_max": -0.2,
        "prob_limpieza_picos": 0.7
    }
}

# --- FUNCIONES DE GENERACIÓN DE CURVAS ---
def generar_deriva_gaussiana(longitud, amplitud, sigma, seed):
    """Genera una curva de deriva suave (aditiva) única por DL."""
    np.random.seed(seed)
    try:
        ruido_base = np.random.randn(longitud)
        deriva_suave = gaussian_filter1d(ruido_base, sigma=sigma)
        max_abs = np.max(np.abs(deriva_suave))
        if max_abs > 1e-6: 
            deriva_normalizada = deriva_suave / max_abs
        else: 
            deriva_normalizada = np.zeros(longitud)
        deriva_final = deriva_normalizada * amplitud
        fade_len = min(longitud // 10, int(sigma * 3))
        if fade_len > 1:
            fade_in = np.linspace(0, 1, fade_len)
            deriva_final[:fade_len] *= fade_in
            fade_out = np.linspace(1, 0, fade_len)
            deriva_final[-fade_len:] *= fade_out
        return deriva_final
    except Exception: 
        return np.zeros(longitud)

def generar_curva_multiplicativa(longitud, variacion_percent, punto_pico_frac):
    """Genera una curva de multiplicación que vuelve a 1.0."""
    try:
        factor_max = 1.0 + variacion_percent
        punto_pico_idx = int(longitud * punto_pico_frac)
        if punto_pico_idx <= 0: 
            punto_pico_idx = 1
        if punto_pico_idx >= longitud: 
            punto_pico_idx = longitud - 1
        
        fase_subida = np.linspace(1.0, factor_max, punto_pico_idx)
        fase_bajada = np.linspace(factor_max, 1.0, longitud - punto_pico_idx)
        curva_multi = np.concatenate((fase_subida[:-1], fase_bajada))
        
        if len(curva_multi) != longitud:
            x_original = np.linspace(0, 1, len(curva_multi))
            x_nuevo = np.linspace(0, 1, longitud)
            curva_multi = np.interp(x_nuevo, x_original, curva_multi)
        return curva_multi
    except Exception: 
        return np.ones(longitud)

def aplicar_pipeline_a_columna(datos_np, config_dl, seed):
    """Aplica el pipeline de 4 pasos a una sola columna de datos."""
    longitud_actual = len(datos_np)
    if longitud_actual < 20:
        return datos_np

    # Sellar la aleatoriedad para esta columna específica
    col_seed = seed + hash(config_dl['dl_nombre']) % 1000
    random.seed(col_seed)
    np.random.seed(col_seed)

    # PASO 1: LIMPIEZA DE PICOS (Probabilística)
    if random.random() < config_dl["prob_limpieza_picos"]:
        datos_base = medfilt(datos_np, kernel_size=3)
    else:
        datos_base = datos_np

    # PASO 2: EXTRAPOLACIÓN (Variable por DL)
    curva_multi_dl = generar_curva_multiplicativa(longitud_actual, config_dl["variacion_percent"], config_dl["punto_pico_frac"])
    datos_extrapolados = datos_base * curva_multi_dl
    
    # PASO 3: DERIVA DE REALISMO (Única por DL)
    deriva = generar_deriva_gaussiana(longitud_actual, config_dl["amplitud"], config_dl["sigma"], seed=col_seed + 1)
    datos_con_deriva = datos_extrapolados + deriva
    
    # PASO 4: APLICAR OFFSET BASE (Variable por DL)
    datos_finales = datos_con_deriva + config_dl["offset_base"]
    
    return datos_finales

def leer_datos_crudos_excel(wb_bytes):
    """Lee TODOS los datos crudos del Excel y los almacena en un dict."""
    datos_crudos = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
        for hoja_nombre in wb.sheetnames:
            if any(ignorar in hoja_nombre.strip().upper() for ignorar in HOJAS_A_IGNORAR):
                continue
            
            ws = wb[hoja_nombre]
            datos_hoja = {}
            for col in ws.iter_cols(min_row=1):
                header_value = col[0].value
                if isinstance(header_value, str) and header_value.strip().upper().startswith("DL"):
                    valores = []
                    for cell in col[1:]:
                        if isinstance(cell.value, (int, float)):
                            valores.append(cell.value)
                    
                    if len(valores) > 20:
                        datos_hoja[header_value.strip()] = np.array(valores)
            
            if datos_hoja:
                datos_crudos[hoja_nombre] = datos_hoja
        return datos_crudos
    except Exception as e:
        logger.error(f"Error al leer el archivo Excel: {e}")
        return None

def generar_configuracion_inicial(datos_crudos, config_base, seed_value):
    """Genera el dict de configuración inicial para CADA DL en TODAS las hojas."""
    random.seed(seed_value)
    np.random.seed(seed_value)
    
    config_hojas = {}
    for hoja_nombre, dls in datos_crudos.items():
        config_hoja_actual = {}
        for dl_nombre in dls.keys():
            config_hoja_actual[dl_nombre] = {
                "dl_nombre": dl_nombre,
                "variacion_percent": random.uniform(config_base["variacion_min"], config_base["variacion_max"]),
                "amplitud": config_base["amplitud"],
                "sigma": config_base["sigma"],
                "punto_pico_frac": config_base["punto_pico"],
                "offset_base": random.uniform(config_base["offset_min"], config_base["offset_max"]),
                "prob_limpieza_picos": config_base["prob_limpieza_picos"]
            }
        config_hojas[hoja_nombre] = config_hoja_actual
            
    return config_hojas

def generar_datos_extrapolados_df(datos_crudos_hoja, config_por_dl, seed_value):
    """Genera un DataFrame extrapolado basado en la configuración de cada DL."""
    datos_extrapolados = {}
    for dl_nombre, datos_originales in datos_crudos_hoja.items():
        if dl_nombre in config_por_dl:
            config_dl = config_por_dl[dl_nombre]
            datos_extrapolados[dl_nombre] = aplicar_pipeline_a_columna(datos_originales, config_dl, seed_value)
    
    return pd.DataFrame(dict([(k, pd.Series(v)) for k, v in datos_extrapolados.items()]))

def crear_grafico_matplotlib(df, titulo, limite_max=None, limite_min=None, seccion_seleccionada=None):
    """Crea un gráfico matplotlib con límites opcionales."""
    if df.empty:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.text(0.5, 0.5, 'No hay datos disponibles', ha='center', va='center', transform=ax.transAxes)
        return fig
    
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Graficar cada columna
    for columna in df.columns:
        ax.plot(df.index, df[columna], label=columna, linewidth=2)
    
    # Añadir límites si existen
    if limite_max is not None:
        ax.axhline(y=limite_max, color='red', linestyle='--', alpha=0.7, label='Límite Máx')
    
    if limite_min is not None:
        ax.axhline(y=limite_min, color='red', linestyle='--', alpha=0.7, label='Límite Mín')
    
    # Añadir área de sección seleccionada
    if seccion_seleccionada and 'inicio' in seccion_seleccionada and 'fin' in seccion_seleccionada:
        ax.axvspan(seccion_seleccionada['inicio'], seccion_seleccionada['fin'], 
                  alpha=0.3, color='yellow', label='Sección Editada')
    
    ax.set_title(titulo, fontsize=14, fontweight='bold')
    ax.set_xlabel('Índice de Tiempo')
    ax.set_ylabel('Valor')
    ax.legend()
    ax.grid(True, alpha=0.3)
    
    return fig

# --- FUNCIONES PRINCIPALES PARA GRADIO ---
def procesar_archivo(archivo, preset_seleccionado, seed_value):
    """Función principal que procesa el archivo y genera las visualizaciones."""
    if archivo is None:
        return None, None, "Por favor carga un archivo .xlsm"
    
    try:
        # Leer datos crudos
        datos_crudos = leer_datos_crudos_excel(archivo)
        if not datos_crudos:
            return None, None, "Error: No se pudieron leer datos válidos del archivo"
        
        # Obtener primera hoja disponible
        hojas_disponibles = list(datos_crudos.keys())
        if not hojas_disponibles:
            return None, None, "Error: No se encontraron hojas con datos DL válidos"
        
        hoja_actual = hojas_disponibles[0]
        
        # Generar configuración
        config_base = CONFIGURACION_BASE[preset_seleccionado]
        config_hojas = generar_configuracion_inicial(datos_crudos, config_base, seed_value)
        
        # Generar datos originales y extrapolados
        df_original = pd.DataFrame(datos_crudos[hoja_actual])
        df_extrapolado = generar_datos_extrapolados_df(
            datos_crudos[hoja_actual], 
            config_hojas[hoja_actual], 
            seed_value
        )
        
        # Determinar límites y títulos
        es_temperatura = "%HR" not in hoja_actual.upper() and "HUM" not in hoja_actual.upper()
        
        if es_temperatura:
            titulo_orig = f"Temperatura Original - {hoja_actual}"
            titulo_ext = f"Temperatura Extrapolada (v{seed_value}) - {hoja_actual}"
            limite_min, limite_max = 15, 25
        else:
            titulo_orig = f"Humedad Original - {hoja_actual}"
            titulo_ext = f"Humedad Extrapolada (v{seed_value}) - {hoja_actual}"
            limite_min, limite_max = None, None
        
        # Crear gráficos
        fig_original = crear_grafico_matplotlib(df_original, titulo_orig, limite_max, limite_min)
        fig_extrapolado = crear_grafico_matplotlib(df_extrapolado, titulo_ext, limite_max, limite_min)
        
        mensaje = f"✅ Archivo procesado exitosamente\n• Hoja: {hoja_actual}\n• Preset: {preset_seleccionado}\n• Versión: {seed_value}\n• Sensores: {len(df_original.columns)}"
        
        return fig_original, fig_extrapolado, mensaje
        
    except Exception as e:
        logger.error(f"Error en procesamiento: {e}")
        return None, None, f"❌ Error al procesar archivo: {str(e)}"

def actualizar_curva_individual(archivo, preset_seleccionado, seed_value, 
                               prob_limpieza, variacion_percent, offset_base, 
                               amplitud, sigma, dl_seleccionado):
    """Actualiza una curva individual con parámetros específicos."""
    if archivo is None:
        return None, "Carga un archivo primero"
    
    try:
        datos_crudos = leer_datos_crudos_excel(archivo)
        if not datos_crudos:
            return None, "Error leyendo datos"
        
        hojas_disponibles = list(datos_crudos.keys())
        hoja_actual = hojas_disponibles[0]
        
        # Configuración base
        config_base = CONFIGURACION_BASE[preset_seleccionado]
        config_hojas = generar_configuracion_inicial(datos_crudos, config_base, seed_value)
        
        # Aplicar ajustes individuales al DL seleccionado
        if dl_seleccionado in config_hojas[hoja_actual]:
            config_hojas[hoja_actual][dl_seleccionado].update({
                "prob_limpieza_picos": prob_limpieza,
                "variacion_percent": variacion_percent,
                "offset_base": offset_base,
                "amplitud": amplitud,
                "sigma": sigma
            })
        
        # Generar datos extrapolados
        df_extrapolado = generar_datos_extrapolados_df(
            datos_crudos[hoja_actual], 
            config_hojas[hoja_actual], 
            seed_value
        )
        
        # Crear gráfico actualizado
        es_temperatura = "%HR" not in hoja_actual.upper() and "HUM" not in hoja_actual.upper()
        titulo = f"Curva Ajustada - {dl_seleccionado}"
        
        if es_temperatura:
            limite_min, limite_max = 15, 25
        else:
            limite_min, limite_max = None, None
        
        fig_ajustado = crear_grafico_matplotlib(df_extrapolado[[dl_seleccionado]], titulo, limite_max, limite_min)
        
        mensaje = f"✅ {dl_seleccionado} actualizado\n• Limpieza: {prob_limpieza}\n• Variación: {variacion_percent:.3f}\n• Offset: {offset_base:.2f}"
        
        return fig_ajustado, mensaje
        
    except Exception as e:
        return None, f"Error actualizando curva: {str(e)}"

# --- INTERFAZ GRADIO ---
with gr.Blocks(theme=gr.themes.Soft(), title="Extrapolador Maestro HF") as demo:
    gr.Markdown("# 🔬 Extrapolador Maestro - HF Spaces")
    gr.Markdown("Genera extrapolaciones realistas de datos de sensores con actualización instantánea")
    
    with gr.Row():
        with gr.Column(scale=1):
            # --- PANEL DE CONTROL ---
            gr.Markdown("## 🎛️ Controles Principales")
            
            archivo = gr.File(
                label="1. Subir archivo .xlsm",
                file_types=[".xlsm"],
                type="binary"
            )
            
            preset_seleccionado = gr.Dropdown(
                label="2. Preset de Prueba",
                choices=list(CONFIGURACION_BASE.keys()),
                value="OQ Mapeo"
            )
            
            seed_value = gr.Slider(
                label="3. Versión (Semilla)",
                minimum=1,
                maximum=100,
                value=1,
                step=1
            )
            
            btn_procesar = gr.Button("🔄 Procesar Archivo", variant="primary")
            
            # --- EDITOR INDIVIDUAL ---
            gr.Markdown("## 🎚️ Editor Individual")
            
            dl_seleccionado = gr.Dropdown(
                label="Sensor a Editar",
                choices=[],
                value=None
            )
            
            with gr.Row():
                prob_limpieza = gr.Slider(0.0, 1.0, value=0.9, label="Limpieza Picos")
                variacion_percent = gr.Slider(0.0, 0.2, value=0.02, label="Extrapolación %")
            
            with gr.Row():
                offset_base = gr.Slider(-2.0, 1.0, value=-0.3, label="Offset Base")
                amplitud = gr.Slider(0.0, 1.0, value=0.3, label="Amplitud Deriva")
            
            sigma = gr.Slider(3, 25, value=12, label="Suavidad Deriva")
            
            btn_actualizar_individual = gr.Button("🎯 Actualizar Curva Individual", variant="secondary")
        
        with gr.Column(scale=2):
            # --- ÁREA DE VISUALIZACIÓN ---
            gr.Markdown("## 📊 Visualización")
            
            with gr.Row():
                with gr.Column():
                    gr.Markdown("### Original")
                    grafico_original = gr.Plot(label="Gráfico Original")
                
                with gr.Column():
                    gr.Markdown("### Extrapolado")
                    grafico_extrapolado = gr.Plot(label="Gráfico Extrapolado")
            
            # Gráfico individual ajustado
            gr.Markdown("### Curva Individual Ajustada")
            with gr.Row():
                grafico_individual = gr.Plot(label="Curva Individual")
                mensaje_individual = gr.Textbox(label="Estado", interactive=False)
            
            # Mensajes de estado
            mensaje_estado = gr.Textbox(
                label="📝 Estado del Procesamiento",
                interactive=False,
                lines=4
            )
    
    # --- EVENTOS Y CONEXIONES ---
    def actualizar_dl_disponibles(archivo):
        """Actualiza la lista de DL disponibles cuando se carga un archivo."""
        if archivo is None:
            return gr.Dropdown(choices=[])
        
        try:
            datos_crudos = leer_datos_crudos_excel(archivo)
            if datos_crudos and len(datos_crudos) > 0:
                primera_hoja = list(datos_crudos.keys())[0]
                dl_disponibles = list(datos_crudos[primera_hoja].keys())
                return gr.Dropdown(choices=dl_disponibles, value=dl_disponibles[0] if dl_disponibles else None)
        except Exception:
            pass
        
        return gr.Dropdown(choices=[])
    
    # Conectar eventos
    archivo.change(
        fn=actualizar_dl_disponibles,
        inputs=[archivo],
        outputs=[dl_seleccionado]
    )
    
    btn_procesar.click(
        fn=procesar_archivo,
        inputs=[archivo, preset_seleccionado, seed_value],
        outputs=[grafico_original, grafico_extrapolado, mensaje_estado]
    )
    
    btn_actualizar_individual.click(
        fn=actualizar_curva_individual,
        inputs=[
            archivo, preset_seleccionado, seed_value,
            prob_limpieza, variacion_percent, offset_base,
            amplitud, sigma, dl_seleccionado
        ],
        outputs=[grafico_individual, mensaje_individual]
    )
    
    # Actualización automática cuando cambian los parámetros individuales
    inputs_individuales = [prob_limpieza, variacion_percent, offset_base, amplitud, sigma]
    
    for input_comp in inputs_individuales:
        input_comp.change(
            fn=actualizar_curva_individual,
            inputs=[
                archivo, preset_seleccionado, seed_value,
                prob_limpieza, variacion_percent, offset_base,
                amplitud, sigma, dl_seleccionado
            ],
            outputs=[grafico_individual, mensaje_individual]
        )
    
    # Información adicional
    gr.Markdown("---")
    gr.Markdown("""
    ### 🚀 Características:
    - **Actualización instantánea** al mover controles
    - **Múltiples presets** de configuración
    - **Editor individual** por sensor
    - **Visualización en tiempo real**
    - **Completamente gratuito** en Hugging Face Spaces
    
    ### 📁 Uso:
    1. Sube tu archivo .xlsm
    2. Selecciona preset y versión
    3. Haz clic en "Procesar Archivo"
    4. Ajusta parámetros individuales
    5. ¡Los cambios se ven al instante!
    """)

if __name__ == "__main__":
    demo.launch(
        share=False,
        show_error=True
    )
